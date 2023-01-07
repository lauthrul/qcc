# 使用openpyxl追加写入数据到Excel
import openpyxl
import os


def open_excel(path: str, sheet_name: str = 'Sheet1'):
    if not os.path.exists(path):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = sheet_name
        wb.save(path)
        return wb
    else:
        wb = openpyxl.load_workbook(path)
        wb.active = wb[sheet_name]
        wb.save(path)  # 尝试保存，如果抛出异常就说明被占用
        return wb


def truncate_sheet(wb: openpyxl.Workbook, sheet_name: str):
    sheet_idx = wb.sheetnames.index(sheet_name)
    sheet = wb[sheet_name]
    wb.remove(sheet)
    wb.create_sheet(sheet_name, sheet_idx)


def append_excel(wb: openpyxl.Workbook, sheet_name: str, value):
    sheet = wb[sheet_name]
    start_row = sheet.max_row  # 获得最大行数
    index = len(value)
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(start_row + i + 1, j + 1, value[i][j])


def read_excel(path: str, sheet_name: str):
    wb = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = wb[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()


def remove_duplicates(path: str, sheet_name: str, cols: list[int]):
    arr = os.path.splitext(path)
    path2 = arr[0] + '[去重]' + arr[1]
    try:
        wb2 = open_excel(path2, sheet_name)
    except:
        print('文件"{}"已在其他程序中打开，请关闭后再重试！'.format(path2))
        exit(0)

    wb = openpyxl.load_workbook(path)
    sheet = wb[sheet_name]
    value_hash = {}
    values = []
    max_col = 0
    for col in cols:
        max_col = max(col, max_col)
    for row in sheet.rows:
        if len(row) > max_col:
            key = ''
            item = []
            for col in cols:
                key += '_' + str(row[col].value) if row[col].value is not None else ''
            for col in row:
                item.append(col.value)
            if value_hash.get(key) != 1:
                values.append(item)
                value_hash[key] = 1

    truncate_sheet(wb2, sheet_name)
    append_excel(wb2, sheet_name, values)
    wb2.save(path2)
    print('[OUT]: {}'.format(path2))


if __name__ == '__main__':
    book_name = 'xlsx格式测试工作簿.xlsx'
    sheet_name = 'xlsx格式测试表'
    values = [
        ["111", "女", "66", "石家庄", "运维工程师"],
        ["222", "男", "55", "南京", "饭店老板"],
        ["333", "女", "27", "苏州", "保安"],
    ]

    try:
        wb = open_excel(book_name, sheet_name)
    except:
        print('文件已在其他程序中打开，请关闭后再重试！')
        exit(0)

    truncate_sheet(wb, sheet_name)
    append_excel(wb, sheet_name, values)
    wb.save(book_name)

    read_excel(book_name, sheet_name)
